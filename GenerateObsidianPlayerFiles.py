import pandas as pd
import os
from openpyxl.worksheet.table import Table
from openpyxl.utils import get_column_letter
import re

SchedulePath = "InputS23/RSC_Schedule.xlsx"
RPVPath = "InputS23/RSC_RPV.xlsx"
SBVPath = "InputS23/RSC_SBV.xlsx"
IDRPath = "InputS23/RSC_IDR.xlsx"

# Check if all Stats Files Exist
if not os.path.exists(SchedulePath):
    print(
        "Schedule File is not Found. Please place it in the Input folder under the name 'RSC_Schedule.xlsx'"
    )
    exit()

if not os.path.exists(RPVPath):
    print(
        "RPV File is not Found. Please place it in the Input folder under the name 'RSC_RPV.xlsx'"
    )
    exit()

if not os.path.exists(SBVPath):
    print(
        "SBV File is not Found. Please place it in the Input folder under the name 'RSC_SBV.xlsx'"
    )
    exit()

if not os.path.exists(IDRPath):
    print(
        "IDR File is not Found. Please place it in the Input folder under the name 'RSC_IDR.xlsx'"
    )
    exit()

print("Loading Excel Sheets into Memory...")

#AtoL = 12 # 12 Columns
#TeamsPerRow = AtoL - 3

# Load the Excel Sheets
ScheduleSheet = pd.ExcelFile(SchedulePath)
RPVSheet = pd.ExcelFile(RPVPath)
SBVSheet = pd.ExcelFile(SBVPath)
IDRSheet = pd.ExcelFile(IDRPath)

Tiers = [
    "Premier",
    "Master",
    "Elite",
    "Veteran",
    "Rival",
    "Challenger",
    "Prospect",
    "Contender",
    "Amateur",
]

schedulePages = ScheduleSheet.sheet_names

MatchDayScheduleSheetName = schedulePages[0]

print("Extracting Match Day Schedules...")

#TierInfo = []
#TierInfo.append({"matchSheetName": schedulePages[0], "scrimSheetName" : schedulePages[1], "skiprowsMatch": 5, "usecolsMatch": "B:I", "nrowsMatch": 49 - 6,       "skiprowsScrim": 13, "usecolsScrim": "B:L", "nrowsScrim": 49-14, "preseasonDays": 3, "matchDays": 16, "playoffDays": 2, "spacing": 4, "matchDaysScrim": 16, "spacingScrim" : 2, "hasLunarConference": True, "hasDivisions": False})
#TierInfo.append({"matchSheetName": schedulePages[0], "scrimSheetName" : schedulePages[2], "skiprowsMatch": 56, "usecolsMatch": "B:K", "nrowsMatch": 102 - 57,    "skiprowsScrim": 13, "usecolsScrim": "B:L", "nrowsScrim": 49-14,  "preseasonDays": 3,"matchDays": 17, "playoffDays": 2,"spacing": 4,  "matchDaysScrim": 16, "spacingScrim" : 2, "hasLunarConference": True, "hasDivisions": False})
#TierInfo.append({"matchSheetName": schedulePages[0], "scrimSheetName" : schedulePages[3], "skiprowsMatch": 109, "usecolsMatch": "B:Q", "nrowsMatch": 158 - 110,  "skiprowsScrim": 13, "usecolsScrim": "B:L", "nrowsScrim": 87-14, "preseasonDays": 3, "matchDays": 18, "playoffDays": 3, "spacing": 5, "matchDaysScrim": 16, "spacingScrim" : 2, "hasLunarConference": True, "hasDivisions": False})
#TierInfo.append({"matchSheetName": schedulePages[0], "scrimSheetName" : schedulePages[4], "skiprowsMatch": 167, "usecolsMatch": "B:V", "nrowsMatch": 218 - 168,  "skiprowsScrim": 13, "usecolsScrim": "B:L", "nrowsScrim": 87-14, "preseasonDays": 3, "matchDays": 19, "playoffDays": 3, "spacing": 5, "matchDaysScrim": 16, "spacingScrim" : 2, "hasLunarConference": True, "hasDivisions": True})
#TierInfo.append({"matchSheetName": schedulePages[0], "scrimSheetName" : schedulePages[5], "skiprowsMatch": 227, "usecolsMatch": "B:V", "nrowsMatch": 278 - 228,  "skiprowsScrim": 13, "usecolsScrim": "B:L", "nrowsScrim": 87-14, "preseasonDays": 3, "matchDays": 19, "playoffDays": 3, "spacing": 5, "matchDaysScrim": 16, "spacingScrim" : 2, "hasLunarConference": True, "hasDivisions": True})
#TierInfo.append({"matchSheetName": schedulePages[0], "scrimSheetName" : schedulePages[6], "skiprowsMatch": 285, "usecolsMatch": "B:Q", "nrowsMatch": 334 - 286,  "skiprowsScrim": 13, "usecolsScrim": "B:L", "nrowsScrim": 87-14, "preseasonDays": 3, "matchDays": 18, "playoffDays": 3, "spacing": 5, "matchDaysScrim": 16, "spacingScrim" : 2, "hasLunarConference": True, "hasDivisions": False})
#TierInfo.append({"matchSheetName": schedulePages[0], "scrimSheetName" : schedulePages[7], "skiprowsMatch": 341, "usecolsMatch": "B:K", "nrowsMatch": 387 - 342,  "skiprowsScrim": 13, "usecolsScrim": "B:L", "nrowsScrim": 49-14, "preseasonDays": 3, "matchDays": 17, "playoffDays": 2, "spacing": 4, "matchDaysScrim": 16, "spacingScrim" : 2, "hasLunarConference": True, "hasDivisions": False})
#TierInfo.append({"matchSheetName": schedulePages[0], "scrimSheetName" : schedulePages[8], "skiprowsMatch": 394, "usecolsMatch": "B:H", "nrowsMatch": 438 - 395,  "skiprowsScrim": 13, "usecolsScrim": "B:L", "nrowsScrim": 49-14, "preseasonDays": 3, "matchDays": 16, "playoffDays": 3, "spacing": 4, "matchDaysScrim": 16, "spacingScrim" : 2, "hasLunarConference": True, "hasDivisions": False})
#TierInfo.append({"matchSheetName": schedulePages[0], "scrimSheetName" : schedulePages[9], "skiprowsMatch": 445, "usecolsMatch": "B:K", "nrowsMatch": 464 - 446,  "skiprowsScrim": 13, "usecolsScrim": "B:k", "nrowsScrim": 30-14, "preseasonDays": 3, "matchDays": 15, "playoffDays": 1, "spacing": 4, "matchDaysScrim": 16, "spacingScrim" : 2, "hasLunarConference": False, "hasDivisions": False})

NumberOfPlayers = 841

StatsInfo = []
StatsInfo.append({"sheet" : RPVSheet, "sheetName": "Sorted RPV", "skiprows": 4, "usecols": "A:AD", "nrows": NumberOfPlayers})
StatsInfo.append({"sheet" : SBVSheet,"sheetName": "Sorted Player Stats", "skiprows": 0, "usecols": "A:AR", "nrows": NumberOfPlayers})
StatsInfo.append({"sheet" : IDRSheet,"sheetName": "IDR View", "skiprows": 4, "usecols": "A:AQ", "nrows": NumberOfPlayers})

TeamsInfoCoords = { "sheet" : SBVSheet, "sheetName" : "Sorted Team Stats", "skiprows": 0, "usecols": "A:AK", "nrows": 177 }

MatchScheduleSheets = {}
ScrimScheduleSheets = {}

print("Excel Sheets Loaded Successfully!")


#for i in range(len(Tiers)):
#    MatchScheduleSheets[Tiers[i]] = LinearizeMatchDaySchedule(ScheduleSheet, TierInfo[i])
#    
#    divisionOffset = 3 if TierInfo[i]["hasDivisions"] else 0
#    conferenceMultiple = 2 if TierInfo[i]["hasLunarConference"] else 1
#    
#    teamColumnNums = len(GetTable(ScheduleSheet, TierInfo[i]["matchSheetName"], TierInfo[i]["skiprowsMatch"], TierInfo[i]["usecolsMatch"], TierInfo[i]["nrowsMatch"]).columns)
#    
#    TierInfo[i]["teamNum"] = (teamColumnNums - 2 - divisionOffset) * conferenceMultiple
#    TierInfo[i]["scrimRows"] = int(TierInfo[i]["teamNum"] / TeamsPerRow) + 1
#    
#    ScrimScheduleSheets[Tiers[i]] = LinearizeScrimDaySchedule(ScheduleSheet, TierInfo[i])


#
# Function Definitions
#

def GetTable(sheet: pd.ExcelFile, sheetName: str, skipRows: int, useCols: str, nRows: int) -> pd.DataFrame:
    return sheet.parse(
        sheetName,
        skiprows=skipRows,
        usecols=useCols,
        nrows=nRows,
    )

def LinearizeMatchDaySchedule (scheduleSheet,  tierInfo):
    
    sheet = GetTable(scheduleSheet, tierInfo["matchSheetName"], tierInfo["skiprowsMatch"], tierInfo["usecolsMatch"], tierInfo["nrowsMatch"])
    
    preseasonDays = tierInfo["preseasonDays"]
    matchDays = tierInfo["matchDays"]
    playoffDays = tierInfo["playoffDays"]
    spacing = tierInfo["spacing"]
    hasLunarConference = tierInfo["hasLunarConference"]
    
    TotalDays = preseasonDays + matchDays
    
    OfficialSchedule = sheet.iloc[:TotalDays, [0, -1]]
    OfficialSchedule.columns = ["Match Day", "Date"]
    
    SolarConference = sheet.iloc[:TotalDays, 1:-1]
    SolarConference = SolarConference.dropna(axis=1, how='all')
    
    # Fill in Playoff Days
    for i in range(TotalDays - playoffDays, TotalDays):
        DayValue = SolarConference.iloc[i, 0]
        SolarConference.iloc[i, :] = DayValue
        
    OfficialSchedule = pd.concat([OfficialSchedule, SolarConference], axis=1)

    if hasLunarConference:
        LunarConferenceTeamNames = sheet.iloc[TotalDays + spacing, 1:-1].values
        LunarConference = sheet.iloc[TotalDays + spacing+1:TotalDays + spacing + TotalDays + 1, 1:-1]
        LunarConference.columns = LunarConferenceTeamNames
        LunarConference.reset_index(drop=True, inplace=True)
        LunarConference = LunarConference.dropna(axis=1, how='all')
        
        # Fill in Playoff Days
        for i in range(TotalDays - playoffDays, TotalDays):
            DayValue = LunarConference.iloc[i, 0]
            LunarConference.iloc[i, :] = DayValue
            
        OfficialSchedule = pd.concat([OfficialSchedule, LunarConference], axis=1)
        OfficialSchedule = OfficialSchedule.dropna(axis=1, how='all')

    return OfficialSchedule

def LinearizeScrimDaySchedule (scheduleSheet, tierInfo):
    
    sheet = GetTable(scheduleSheet, tierInfo["scrimSheetName"], tierInfo["skiprowsScrim"], tierInfo["usecolsScrim"], tierInfo["nrowsScrim"])
    
    numOfRows = tierInfo["scrimRows"]
    matchDays = tierInfo["matchDaysScrim"]
    spacing = tierInfo["spacingScrim"]
    
    OfficialSheet = sheet.iloc[:matchDays, [0, 2]]
    FirstRow = sheet.iloc[:matchDays, 3:]
    OfficialSheet = pd.concat([OfficialSheet, FirstRow], axis=1)
    
    Start = matchDays + spacing
    
    for i in range(1, numOfRows):
        RowNames = sheet.iloc[Start][2:].values
        Row = sheet.iloc[(Start + 1):(Start + 1) + matchDays, 2:]
        Row.columns = RowNames
        Row.reset_index(drop=True, inplace=True)
        Row.dropna(axis=1, how='all')
        Row.dropna(axis=0, how='all')
        
        Start += (matchDays + spacing) 
        
        OfficialSheet = pd.concat([OfficialSheet, Row], axis=1)
    
    OfficialSheet = OfficialSheet.loc[:, ~OfficialSheet.columns.isna()]
    ScrimNumbers = len(OfficialSheet["Game"].values)
    OfficialSheet["Game"] =  ["Match Day " + str(i) for i in range(1, ScrimNumbers + 1)]
    OfficialSheet.rename(columns={"Game": "Match Day"}, inplace=True)
    
    return OfficialSheet

def GetStatsSheet (statsInfo):
    
    RPVTable = GetTable(statsInfo[0]["sheet"], statsInfo[0]["sheetName"], statsInfo[0]["skiprows"], statsInfo[0]["usecols"], statsInfo[0]["nrows"])
    SBVTable = GetTable(statsInfo[1]["sheet"], statsInfo[1]["sheetName"], statsInfo[1]["skiprows"], statsInfo[1]["usecols"], statsInfo[1]["nrows"])
    IDRTable = GetTable(statsInfo[2]["sheet"], statsInfo[2]["sheetName"], statsInfo[2]["skiprows"], statsInfo[2]["usecols"], statsInfo[2]["nrows"])
    
    SBVTable.rename(columns={"Player Name": "Name", "Team(s)": "Team", "Conf." : "Conference"}, inplace=True)
    IDRTable.rename(columns={"Team": "Current Team"}, inplace=True)
    
    mergedTable = (
            RPVTable
            .merge(SBVTable, on=["Name"])
            .merge(IDRTable, on=["Name"])
            )
    
    mergedTable = mergedTable.rename(columns={col: col.replace("_x", "") for col in mergedTable.columns if "_x" in col})
    mergedTable = mergedTable.rename(columns={col: col.replace("(s)", "_y") for col in mergedTable.columns if "(s)" in col})

    # Remove columns containing "_y" or "_z"
    remove_cols = [col for col in mergedTable.columns if "_y" in col or "_z" in col]
    mergedTable = mergedTable.drop(columns=remove_cols)
    
    mergedTable = mergedTable.loc[:, ~mergedTable.columns.duplicated()]
    
    return mergedTable
    
def GetExcelRange(start_row, start_col, df):
    rows, cols = df.shape
    end_row = start_row + rows  # Bottom-right row
    end_col = start_col + cols - 1  # Bottom-right column (adjust for zero indexing)
    
    
    start_cell = f"{get_column_letter(start_col)}{start_row}"
    end_cell = f"{get_column_letter(end_col)}{end_row}" # Index = False, Headers = True, adjust based off that
    
    return f"{start_cell}:{end_cell}"

def PasteDataFrame(dataFrame, workSheet, startRow = 1, startCol = 1):
    for j, col_name in enumerate(dataFrame.columns, start=startCol):  # Write column headers
        workSheet.cell(row=startRow, column=j).value = col_name
    
    for row_num, row_data in enumerate(dataFrame.itertuples(index=False), start=startRow + 1):  # Write data
        for col_num, value in enumerate(row_data, start=startCol):
            workSheet.cell(row=row_num, column=col_num).value = value

def AddDataFrameAsTable (dataFrame, workSheet, tableName, startRow = 1, startCol = 1):
    PasteDataFrame(dataFrame, workSheet, startRow, startCol)

    if (tableName not in workSheet.tables):
        workSheet.add_table(Table(displayName=tableName, ref=GetExcelRange(startRow, startCol, dataFrame)))

def SanitizeFileName(filename, replacement="_"):
    # Define invalid characters for Windows file names
    invalid_chars = r'[\\/:*?"<>|]'
    # Replace them with the given replacement character (default: "_")
    return re.sub(invalid_chars, replacement, filename)

def GetTeamSheet (teamInfo):
    TeamSheet = teamInfo["sheet"].parse(teamInfo["sheetName"], skiprows=teamInfo["skiprows"], usecols=teamInfo["usecols"], nrows=teamInfo["nrows"])
    TeamSheet.rename(columns={"Conf": "Conference"}, inplace=True)
    return TeamSheet

#
# Obsidian File Generation Functions
#

def GenerateObsidianTierFiles ():
    
    os.makedirs(f"Output/Tiers", exist_ok=True)
    
    rank = 1
    
    for i in range(len(Tiers)):
        Tier = Tiers[i]
        
        obsidianFileName = SanitizeFileName(Tier) + ".md"
        
        obsidianFileContent = "---\n"
        obsidianFileContent += f"Tier: {Tier}\n"
        obsidianFileContent += f"Tier Rank: {rank}\n"
        obsidianFileContent += "---\n"
        
        rank += 1
        
        with open(f"Output/Tiers/{obsidianFileName}", "w") as file:
            file.write(obsidianFileContent)

def GenerateObsidianFranchiseFiles ():
        
    os.makedirs(f"Output/Franchises", exist_ok=True)
    
    franchises = TeamInfoTable["Franchise"].unique()
    
    for franchise in franchises:
        
        links = []
        teamsOnFranchise = TeamInfoTable[TeamInfoTable["Franchise"] == franchise]
        
        links.append(teamsOnFranchise["Tier"].values)
        links.append(teamsOnFranchise["Team"].values)
        
        obsidianFileName = SanitizeFileName(franchise) + ".md"
        
        obsidianFileContent = "---\n"
        obsidianFileContent += f"Franchise: {franchise}\n"
        
        obsidianFileContent += "Ranks:\n"
        for tier in teamsOnFranchise["Tier"].values:
            obsidianFileContent += f"  - \"[[{tier}]]\"\n"
            
        obsidianFileContent += "Teams:\n"
        for tier in teamsOnFranchise["Team"].values:
            obsidianFileContent += f"  - \"[[{tier}]]\"\n"
        
        obsidianFileContent += "---\n"
        
        obsidianFileContent += f"# {franchise}\n---\n"
        obsidianFileContent += f"## Teams\n---\n"
        
        for i in range(len(links[0])):
            obsidianFileContent += f"[[{links[1][i]}]] ([[{links[0][i]}]])\n"
        
        with open(f"Output/Franchises/{obsidianFileName}", "w") as file:
            file.write(obsidianFileContent)
    
def GenerateObsidianTeamFiles ():
        
    os.makedirs(f"Output/Teams", exist_ok=True)
    
    nonLinkItems = ["Team", "Conference", "SBV"]
    linkItems = ["Tier", "Franchise"]
    
    stat_labels = {
    "WP": "Win Percentage",
    "W": "Wins",
    "L": "Losses",
    "P": "Points",
    "SP": "Standard Points",
    "NP": "Non-Standard Points",
    "G": "Goals",
    "A": "Assists",
    "Sv": "Saves",
    "S": "Shots",
    "S%": "Shot Percentage",
    "P/A": "Points per Average",
    "SP/A": "Standard Points per Average",
    "NP/A": "Non-Standard Points per Average",
    "G/A": "Goals per Average",
    "A/A": "Assists per Average",
    "Sv/A": "Saves per Average",
    "S/A": "Shots per Average",
    "P/G": "Points per Game",
    "SP/G": "Standard Points per Game",
    "NP/G": "Non-Standard Points per Game",
    "G/G": "Goals per Game",
    "A/G": "Assists per Game",
    "Sv/G": "Saves per Game",
    "S/G": "Shots per Game",
    "opp_G": "Opponent Goals",
    "opp_G/A": "Opponent Goals per Average",
    "opp_G/G": "Opponent Goals per Game"
}
    
    print(TeamInfoTable)
    
    for index, teamRow in TeamInfoTable.iterrows():
        team = teamRow["Team"]
        
        obsidianFileName = SanitizeFileName(team) + ".md"
        
        obsidianFileContent = "---\n"
        
        for item in linkItems:
            obsidianFileContent += f"{item}: \"[[{teamRow[item].replace("\n", "")}]]\"\n"
        
        for item in nonLinkItems:
            obsidianFileContent += f"{item}: {teamRow[item]}\n"
            
        for item in stat_labels.keys():
            obsidianFileContent += f"{stat_labels[item]}: {teamRow[item]}\n"
        
        obsidianFileContent += "---\n"
    
        with open(f"Output/Teams/{obsidianFileName}", "w") as file:
            file.write(obsidianFileContent)

def GenerateObsidianPlayerFiles ():
    
    unlinkedItems = ["Name", "RSC ID", "Tier", "SBV", "IDR", "RPV"]
    linkedItems = ["Tier"]
    
    stat_labels = {
    "Win%": "WinPercentage",
    "W": "Wins",
    "L": "Losses",
    "P": "Points",
    "SP": "Standard Points",
    "NP": "Non-Standard Points",
    "G": "Goals",
    "A": "Assists",
    "Sv": "Saves",
    "Sh": "Shots",
    "Sh%": "Shot Percentage",
    "P/A": "Points per Average",
    "SP/A": "Standard Points per Average",
    "NP/A": "Non-Standard Points per Average",
    "G/A": "Goals per Average",
    "A/A": "Assists per Average",
    "Sv/A": "Saves per Average",
    "S/A": "Shots per Average",
    "PPG": "Points per Game",
    "SPPG": "Standard Points per Game",
    "NPPG": "Non-Standard Points per Game",
    "GPG": "Goals per Game",
    "APG": "Assists per Game",
    "SvPG": "Saves per Game",
    "ShPG": "Shots per Game",
    "tm_G": "Team Goals",
    "opp_G": "Opponent Goals",
    "opp_G/A": "Opponent Goals per Average",
    "opp_G/G": "Opponent Goals per Game",
    "Current Team": "Team"
    }
    
    os.makedirs(f"Output/Players", exist_ok=True)
    
    for index, playerRow in StatsSheet.iterrows():
        
        playerName = SanitizeFileName(playerRow["Name"])
        obsidianFileName = playerName + ".md"
        
        obsidianFileContent = "---\n"
        
        obsidianFileContent += f"Teams:\n"
        
        for team in playerRow["Team"].split("\n"):
            obsidianFileContent += f"  - \"[[{SanitizeFileName(team)}]]\"\n"
            
        if (isinstance(playerRow["Conference"], float)):
            obsidianFileContent += f"Conference: None\n"
        else:
            obsidianFileContent += f"Conference:\n"
            
            for team in playerRow["Conference"].split("\n"):
                obsidianFileContent += f"  - \"{team}\"\n"
        
        for item in linkedItems:
            obsidianFileContent += f"{item}: [[{playerRow[item].replace("\n", "")}]]\n"
            
        for item in unlinkedItems:
            obsidianFileContent += f"{item}: {playerRow[item]}\n"
            
        for item in stat_labels.keys():
            obsidianFileContent += f"{stat_labels[item]}: {playerRow[item]}\n"
        
        obsidianFileContent += "---\n"
        
        with open(f"Output/Players/{obsidianFileName}", "w") as file:
            file.write(obsidianFileContent)

print("Match Day Schedules Extracted Successfully!")

print("Extracting Stats Data...")

StatsSheet = GetStatsSheet(StatsInfo)

print("Stats Data Extracted Successfully!")
print("Extracting Teams Info...")

TeamInfoTable = GetTeamSheet(TeamsInfoCoords)

print("Teams Info Extracted Successfully!")
print(StatsSheet)
print(StatsSheet.columns)
print(TeamInfoTable)

GenerateObsidianTierFiles()
GenerateObsidianFranchiseFiles()
GenerateObsidianTeamFiles()
GenerateObsidianPlayerFiles()

print("Obsidian Files Generated Successfully!")